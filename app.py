import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

# 1. Configuración de la página web
st.set_page_config(page_title="Mi Excel Ligero", page_icon="📊", layout="wide")
st.title("📊 Generador de tablas personal")
st.write("Gestiona tus columnas arriba, edita tus datos abajo y personaliza el diseño a la izquierda.")

# 2. Inicializar la tabla en la memoria de la sesión si no existe
if "df_datos" not in st.session_state:
    data_inicial = {
        "Columna A": ["", "", "", "", ""],
        "Columna B": ["", "", "", "", ""],
        "Columna C": ["", "", "", "", ""]
    }
    st.session_state.df_datos = pd.DataFrame(data_inicial)

# 3. Barra lateral limpia: Solo para el diseño general
st.sidebar.header("🎨 Configuración del Diseño")
titulo_tabla = st.sidebar.text_input("📝 Título superior de la tabla", placeholder="Ej. Reporte de Ventas")
color_elegido = st.sidebar.color_picker("Color de fondo del encabezado", "#365F91")
color_hex = color_elegido.replace("#", "")

tamano_letra_datos = st.sidebar.slider("Tamaño de letra de los datos", min_value=10, max_value=20, value=12)
alineacion = st.sidebar.selectbox("Alineación del texto", ["Centrado", "Izquierda", "Derecha"])

dict_alineacion = {
    "Centrado": Alignment(horizontal="center", vertical="center"),
    "Izquierda": Alignment(horizontal="left", vertical="center"),
    "Derecha": Alignment(horizontal="right", vertical="center")
}
alineacion_final = dict_alineacion[alineacion]


# 4. Zona Principal: Gestión de Columnas (¡Adiós barra lateral!)
st.subheader("🛠️ Gestión de Columnas")

# Fila para agregar una nueva columna
col_input, col_btn = st.columns([3, 1])
with col_input:
    nueva_columna = st.text_input("Nombre para agregar nueva columna", placeholder="Ej. Total", label_visibility="collapsed")
with col_btn:
    if st.button("➕ Agregar Columna", use_container_width=True):
        if nueva_columna:
            if nueva_columna not in st.session_state.df_datos.columns:
                st.session_state.df_datos[nueva_columna] = [""] * len(st.session_state.df_datos)
                st.rerun()
            else:
                st.error("Esa columna ya existe.")
        else:
            st.warning("Escribe un nombre.")

st.write("**Haz clic en el botón de cualquier columna para cambiarle el nombre:**")

# ✨ EL TRUCO: Creamos una fila de botones horizontal, uno para cada columna actual
columnas_actuales = list(st.session_state.df_datos.columns)
bloques_columnas = st.columns(len(columnas_actuales))

for i, nombre_col in enumerate(columnas_actuales):
    with bloques_columnas[i]:
        # Usamos un componente 'popover' que actúa como un botón que abre una cajita al presionarlo
        with st.popover(f"✏️ {nombre_col}", use_container_width=True):
            nuevo_nombre = st.text_input(f"Nuevo nombre para '{nombre_col}':", key=f"input_renombrar_{i}")
            if st.button("Aplicar", key=f"btn_renombrar_{i}"):
                if nuevo_nombre:
                    if nuevo_nombre not in st.session_state.df_datos.columns:
                        # Renombramos la columna en el DataFrame
                        st.session_state.df_datos = st.session_state.df_datos.rename(columns={nombre_col: nuevo_nombre})
                        st.success("¡Cambiado!")
                        st.rerun()
                    else:
                        st.error("Ya existe ese nombre.")
                else:
                    st.warning("Escribe un nombre válido.")

st.markdown("---")

# 5. Edición de Datos dentro del Formulario
st.subheader("📝 Datos de la Tabla")

with st.form("contenedor_tabla"):
    tabla_editada = st.data_editor(
        st.session_state.df_datos, 
        num_rows="dynamic", 
        use_container_width=True
    )
    
    boton_guardar = st.form_submit_button("💾 Guardar cambios en la tabla")
    
    if boton_guardar:
        st.session_state.df_datos = tabla_editada
        st.success("¡Datos guardados con éxito!")

# 6. Función para procesar el archivo Excel
def generar_excel(dataframe, color_bg, tamano_fuente, alineacion_obj, texto_titulo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    columnas = list(dataframe.columns)
    fila_inicio_tabla = 1
    
    if texto_titulo:
        ws["A1"] = texto_titulo
        ws["A1"].font = Font(name="Arial", size=16, bold=True, color="000000")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        
        if len(columnas) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
        
        fila_inicio_tabla = 3

    for col_num, nombre_columna in enumerate(columnas, start=1):
        ws.cell(row=fila_inicio_tabla, column=col_num, value=nombre_columna)

    for index, fila in dataframe.iterrows():
        fila_excel = fila_inicio_tabla + 1 + index
        for col_num, valor in enumerate(list(fila), start=1):
            ws.cell(row=fila_excel, column=col_num, value=valor)

    fill_encabezado = PatternFill(start_color=color_bg, end_color=color_bg, fill_type="solid")
    fuente_encabezado = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    fuente_datos = Font(name="Arial", size=tamano_fuente, bold=False)

    total_filas = fila_inicio_tabla + len(dataframe)
    for num_fila in range(fila_inicio_tabla, total_filas + 1):
        for num_col in range(1, len(columnas) + 1):
            celda = ws.cell(row=num_fila, column=num_col)
            celda.alignment = alineacion_obj
            
            if num_fila == fila_inicio_tabla:
                celda.fill = fill_encabezado
                celda.font = fuente_encabezado
            else:
                celda.font = fuente_datos

    for col_index, col in enumerate(ws.columns, start=1):
        max_len = 0
        col_letter = get_column_letter(col_index)
        
        for celda in col:
            if celda.row == 1 and texto_titulo:
                continue
            if celda.value is not None:
                len_texto = len(str(celda.value))
                if len_texto > max_len:
                    max_len = len_texto
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# 7. Botón de descarga
st.markdown("---")
if st.button("🚀 Generar y preparar descarga de Excel"):
    archivo_excel = generar_excel(st.session_state.df_datos, color_hex, tamano_letra_datos, alineacion_final, titulo_tabla)
    
    st.download_button(
        label="📥 Descargar archivo .xlsx",
        data=archivo_excel,
        file_name="mi_tabla_ligera.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    st.success("¡Tu archivo está listo para descargar!")
